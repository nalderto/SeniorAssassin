import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from tkinter import *
from tkinter.filedialog import askopenfilename
import os
import random
import copy
import re

seniors_array = []
wb = None
ws = None
fromAddress = ""
password = ""

# Tkinter Window Setup
main = Tk()
main.title("Senior Assassin")
main.geometry("530x340")

header_frame = Frame(main)
header_frame.grid(row = 0, column = 1, sticky = W, pady = 2)

# Excel Sheet Label
excel_sheet_label = Label(header_frame, text = "File Name:")
excel_sheet_label.grid(row = 0, column = 0, sticky = W, pady = 2)

# Excel Sheet Name Label
excel_sheet_name_label = Label(header_frame, text = "No File Initialized")
excel_sheet_name_label.grid(row = 0, column = 1, sticky = W, pady = 2)

# Status Label
status_label = Label(main, text = "")
status_label.grid(row = 1, column = 1, sticky = W, pady = 2)

from_address_label = Label(main, text = "From Address")
fromAddress = Entry(main)
from_address_label.grid(row = 2, column = 0, sticky = W, pady = 2)
fromAddress.grid(row = 2, column = 1, sticky="nsew", pady = 2)

password_label = Label(main, text = "Password")
password = Entry(main, show = "*")
password_label.grid(row = 3, column = 0, sticky = W, pady = 2)
password.grid(row = 3, column = 1, sticky="nsew", pady = 2)

# Assignment Scroll Table
assignment_frame = Frame(main)
assignment_frame.grid(row = 4, column = 1, sticky="nsew", pady = 2)

assassin_list = Listbox(assignment_frame, width = 43)
assassin_list.pack(side = "left", fill = "y")
scrollbar = Scrollbar(assignment_frame, orient="vertical")
scrollbar.pack(side = RIGHT, fill=Y)
scrollbar.config(command = assassin_list.yview)
assassin_list.config(yscrollcommand = scrollbar.set)

assassin_list.insert(END, "There is currently no data")


# Function for closing window
def close_window():
    main.destroy()

# openpyxl Workbook Setup
def workbook_setup():
    global seniors_array
    global wb
    global ws
    file_path = askopenfilename()
    extension = os.path.splitext(file_path)[1]
    basename = os.path.basename(file_path)
    if extension ==".xlsx":
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[0]
        if ws.cell(row = 1, column = 2).value == "First Name":
            status_label.config(text="Workbook Successfully Opened")
            excel_sheet_name_label.config(text=basename)
            assassin_list.delete(0, END)
            for row in range(2, ws.max_row + 1):
                current_person = []
                for column in range(2, 5):
                    current_person.append(ws.cell(row=row, column=column).value)
                seniors_array.append(current_person)
            selection_pool = copy.deepcopy(seniors_array)

            for current in range(0, len(seniors_array)):
                random_pick = random.choice(selection_pool)
                while seniors_array[current] == random_pick:
                    random_pick = random.choice(selection_pool)
                seniors_array[current].append(random_pick[0])
                seniors_array[current].append(random_pick[1])
                selection_pool.remove(random_pick)

                #Update Listbox
                assassin_list.insert(END, str(seniors_array[current][0]) + " " + str(seniors_array[current][1]) + " - " + str(seniors_array[current][3]) + " " + str(seniors_array[current][4]) + " - " + str(seniors_array[current][2]))

        else:
            status_label.config(text="Error: Not a Valid File")
            wb = None
            ws = None
    else:
        status_label.config(text="Error: Not a Valid File")
        wb = None
        ws = None
    print(seniors_array)

def save():
    master_wb = openpyxl.Workbook()
    master_ws = master_wb.active
    master_ws.cell(row = 1, column = 1).value = "Assassin First"
    master_ws.cell(row = 1, column = 2).value = "Assassin Last"
    master_ws.cell(row = 1, column = 3).value = "Assassin Email"
    master_ws.cell(row = 1, column = 4).value = "Target First"
    master_ws.cell(row = 1, column = 5).value = "Target Last"
    currentDirectory = os.getcwd()
    for current in range(0, len(seniors_array)):
        row = current + 2
        master_ws.cell(row = row, column = 1).value = seniors_array[current][0]
        master_ws.cell(row = row, column = 2).value = seniors_array[current][1]
        master_ws.cell(row = row, column = 3).value = seniors_array[current][2]
        master_ws.cell(row = row, column = 4).value = seniors_array[current][3]
        master_ws.cell(row = row, column = 5).value = seniors_array[current][4]
        master_wb.save(str(currentDirectory) + "/Master_List.xlsx")

def email():
    if wb == None or ws == None:
        status_label.config(text="Error: You Must Initiate a File")

    if not isEmailValid(fromAddress):
        status_label.config(text="Error: Invalid Formatted From Email Address")

    if password == "":
        status_label.config(text="Error: Password empty")


    else:
        # SMTP Settings
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()

        try:
            server.login(fromAddress, password)
        except:
            status_label.config(text="Error: Unable to authenticate.  Double check your from email address and password.")

        for current in range(0, len(seniors_array)):
            # To, From, Message
            toaddr = seniors_array[current][2]
            msg = MIMEMultipart()
            msg['From'] = fromAddress
            msg['To'] = toaddr
            msg['Subject'] = "Senior Assassin"

            # Message Body
            body = seniors_array[current][0] + " " + seniors_array[current][1]+ ",\n" + "You have been assigned  " + seniors_array[current][3] + " " + seniors_array[current][4]
            msg.attach(MIMEText(body, 'plain'))


            text = msg.as_string()
            server.sendmail(fromAddress, toaddr, text)
            server.quit()

# Determine if email is valid
def isEmailValid(email):
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
    return re.fullmatch(regex, email)

buttons = Frame(main)
buttons.grid(row = 5, column = 1, sticky=E)

# Initiate button
init_button = Button(buttons, text = "Initiate Workbook", command = workbook_setup)
init_button.grid(row = 0, column = 1, padx = 2)

# Email Button
email_button = Button(buttons, text = "Email", command = email)
email_button.grid(row = 0, column = 2, padx = 2)

# Save Button
save_button = Button(buttons, text = "Save", command = save)
save_button.grid(row = 0, column = 3, padx = 2)

# Quit Button
quit_button = Button(buttons, text = "Quit", command = close_window)
quit_button.grid(row = 0, column = 4, padx = 2)

main.mainloop()
